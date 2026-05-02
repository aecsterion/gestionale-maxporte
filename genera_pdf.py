#!/usr/bin/env python3
"""Motore generazione PDF preventivi/ordini Max Porte"""
import sys, json, os, subprocess, re, glob
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'template_preventivo.xlsx')
VERSION = "2026-05-02-v7"  # marker deploy
if not os.path.exists(TEMPLATE_PATH):
    for c in ['/app/template_preventivo.xlsx', os.path.join(os.getcwd(), 'template_preventivo.xlsx')]:
        if os.path.exists(c): TEMPLATE_PATH = c; break

LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'logo-report.png')
if not os.path.exists(LOGO_PATH):
    for c in ['/app/logo-report.png', os.path.join(os.getcwd(), 'logo-report.png')]:
        if os.path.exists(c): LOGO_PATH = c; break

def aggiungi_logo(pdf_path, logo_path):
    """Sovrappone il logo su ogni pagina del PDF."""
    if not os.path.exists(logo_path):
        print(f"Logo non trovato: {logo_path}", file=sys.stderr)
        return
    try:
        from reportlab.pdfgen import canvas as rl_canvas
        from reportlab.lib.units import mm
        from pypdf import PdfReader, PdfWriter
        import io

        reader = PdfReader(pdf_path)
        writer = PdfWriter()

        for page in reader.pages:
            w = float(page.mediabox.width)
            h = float(page.mediabox.height)
            buf = io.BytesIO()
            c = rl_canvas.Canvas(buf, pagesize=(w, h))
            c.drawImage(logo_path, 14*mm, h - 35*mm,
                        width=28*mm, height=28*mm,
                        preserveAspectRatio=True, mask='auto')
            c.save()
            buf.seek(0)
            overlay = PdfReader(buf).pages[0]
            page.merge_page(overlay)
            writer.add_page(page)

        tmp = pdf_path + '_logo'
        with open(tmp, 'wb') as f:
            writer.write(f)
        os.replace(tmp, pdf_path)
    except Exception as e:
        print(f"Logo warning: {e}", file=sys.stderr)

def fmt_euro(v):
    if not v and v != 0: return ''
    try:
        f = float(v)
        if f == 0: return ''
        return f"{f:,.2f}".replace(',','X').replace('.',',').replace('X','.')
    except: return str(v) if v else ''

def sostituisci(ws, mapping):
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell, MergedCell): continue
            if cell.value and isinstance(cell.value, str):
                val = cell.value
                for k, v in mapping.items():
                    val = val.replace(k, str(v) if v is not None else '')
                cell.value = val

# Range fissi della sezione posizione (letti dal template una volta sola)
_POS_RANGES = {}
def _init_pos_ranges():
    global _POS_RANGES
    if _POS_RANGES: return
    wb = load_workbook(TEMPLATE_PATH)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        start = end = None
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell, MergedCell): continue
                if cell.value and '*POSIZIONE*' in str(cell.value):
                    start = cell.row + 1
                if cell.value and '*TOTALE_POSIZIONE*' in str(cell.value):
                    end = cell.row
        _POS_RANGES[sheet] = (start, end)

def pulisci_euro_vuoti(ws, sheet_name=None):
    """Rimuove euro e sconto vuoti dalle righe senza prezzo listino nella sezione posizione."""
    _init_pos_ranges()
    pos_start, pos_end = _POS_RANGES.get(sheet_name or ws.title, (None, None))
    # Solo sulla sezione posizione — non toccare riepilogo o intestazione
    if not pos_start or not pos_end:
        return

    def is_euro_vuoto(v):
        if not v: return False
        return bool(re.match(r'^€\s*$', str(v).strip()))
    def is_solo_pct(v):
        if not v: return False
        return bool(re.match(r'^\d+%$', str(v).strip()))
    def ha_prezzo(v):
        if not v: return False
        s = str(v).strip()
        return bool(re.search(r'\d', s))

    for row in ws.iter_rows(min_row=pos_start, max_row=pos_end):
        cells = {cell.column: cell for cell in row if not isinstance(cell, MergedCell)}
        # Controlla se c'è un prezzo listino valorizzato (col28)
        v28 = cells.get(28)
        prezzo_listino_ok = v28 and ha_prezzo(v28.value if v28 else None)
        if not prezzo_listino_ok:
            for col, cell in cells.items():
                if cell.value and (is_euro_vuoto(cell.value) or is_solo_pct(cell.value)):
                    cell.value = ''

def nascondi_vuote(ws, sheet_name=None):
    def ok(v):
        if not v: return False
        s = str(v).strip()
        if not s or '*' in s: return False
        if re.match(r'^€\s*$', s) or re.match(r'^\d+%$', s): return False
        if s in ('-','X','x','€'): return False
        return True

    _init_pos_ranges()
    pos_start, pos_end = _POS_RANGES.get(sheet_name or ws.title, (None, None))

    # Nascondi righe della sezione posizione con valore vuoto
    if pos_start and pos_end:
        for row in ws.iter_rows(min_row=pos_start, max_row=pos_end):
            cells = {c.column: c for c in row if not isinstance(c, MergedCell)}
            v16 = cells.get(16)
            v28 = cells.get(28)
            desc_ok = ok(v16.value if v16 else None)
            prz_ok = ok(v28.value if v28 else None)
            # Controlla anche col22 (label totale) e col36 (totale posizione)
            v22 = cells.get(22)
            v36 = cells.get(36)
            extra_ok = ok(v22.value if v22 else None) or ok(v36.value if v36 else None)
            # Nascondi solo se sia descrizione che prezzo listino sono vuoti
            if not desc_ok and not prz_ok and not extra_ok:
                ws.row_dimensions[row[0].row].hidden = True

    # Nascondi righe riepilogo opzionali con valore vuoto (col 37)
    # Nascondi righe riepilogo opzionali con valore vuoto (SOLO su PAGINA_FINALE)
    if (sheet_name or ws.title) == 'PAGINA_FINALE':
        righe_opzionali_riepilogo = {33, 34, 35, 39, 40, 41}
        for row_idx in righe_opzionali_riepilogo:
            try:
                row = list(ws.iter_rows(min_row=row_idx, max_row=row_idx))[0]
                cells = {c.column: c for c in row if not isinstance(c, MergedCell)}
                v37 = cells.get(37)
                val = v37.value if v37 else None
                def val_vuoto(v):
                    if not v: return True
                    s = str(v).strip()
                    return not s or '*' in s or re.match(r'^€\s*$', s)
                if val_vuoto(val):
                    ws.row_dimensions[row_idx].hidden = True
            except: pass

def map_riga(r, sc):
    sc = float(sc)
    def s(v): 
        try: return fmt_euro(round(float(v)*(1-sc/100),2)) if v else ''
        except: return ''
    return {
        '*POSIZIONE*': r.get('posizione',''),
        '*L*': str(r.get('larghezza','')),
        '*H*': str(r.get('altezza','')),
        '*S*': str(r.get('spessore','')),
        '*COD_SENSO*': r.get('senso',''),
        '*COD_APERTURA*': r.get('codice_apertura','') or r.get('apertura',''),
        '*SENSO*': r.get('senso',''),
        '*APERTURA*': r.get('nome_apertura', r.get('apertura','')),
        '*QUANTITÀ*': str(r.get('quantita',1)),
        '*UM*': r.get('um','NR'),
        '*PRODOTTO*': r.get('modello',''),
        '*SERIE*': r.get('serie',''),
        '*MODELLO*': r.get('modello',''),
        '*FINITURA*': r.get('finitura',''),
        '*FINITURA_TELAIO*': r.get('finitura_telaio',''),
        '*FINITURA_COPRIFILI*': r.get('finitura_coprifili',''),
        '*COLORE_PIETRA*': r.get('colore_pietra',''),
        '*COLORE_INSERTO_ALLUMINIO*': r.get('colore_inserto',''),
        '*TIPO_DI_VETRO*': r.get('vetro',''),
        '*INCISIONI_O_STAMPE_DECORATIVE*': r.get('incisioni',''),
        '*BUGNA_O_PANNELLO*': r.get('bugna',''),
        '*TIPOLOGIA*': r.get('nome_apertura', r.get('apertura','')),
        '*PREZZO_TIPOLOGIA_LISTINO*': '',
        '*PREZZO_TIPOLOGIA_SCONTATO*': '',
        '*SERRATURA*': r.get('serratura',''),
        '*SPALLA*': r.get('spalla',''),
        '*FERRAMENTA*': r.get('ferramenta',''),
        '*MANIGLIA* - *VERSIONE_MANIGLIA*': '',
        '*MANIGLIA*': r.get('maniglia',''),
        '*VERSIONE_MANIGLIA*': r.get('versione_maniglia',''),
        '*COLORE_MANIGLIA*': r.get('colore_maniglia',''),
        '*KIT_VARSAVIA_SCORREVOLE*': r.get('kit_varsavia',''),
        '*KIT_RIM_16_SCORREVOLE*': r.get('kit_rim16',''),
        '*LAVORAZIONI_EXTRA*': '',
        '*ACCESSORI*': '',
        '*SUPPLEMENTO_FUORI_MISURA_L_SI_NO*': r.get('fuori_misura_l',''),
        '*SUPPLEMENTO_FUORI_MISURA_H_SI_NO*': r.get('fuori_misura_h',''),
        '*SUPPLEMENTO_RIFILATURA_SI_NO*': '',
        '*STANZA*': r.get('stanza',''),
        '*NOTE_RIGA*': r.get('note_riga',''),
        '*SC*': str(int(sc)) if sc else '',  # sconto applicato solo se c'è un prezzo
        '*PREZZO_MODELLO_LISTINO*': fmt_euro(r.get('prezzo_base',0)),
        '*PREZZO_MODELLO_SCONTATO*': s(r.get('prezzo_base',0)),
        '*PREZZO_FINITURA_LISTINO*': fmt_euro(r.get('prezzo_finitura','')) if r.get('prezzo_finitura') else '',
        '*PREZZO_FINITURA_SCONTATO*': s(r.get('prezzo_finitura',0)) if r.get('prezzo_finitura') else '',
        '*SUPPLEMENTO_BICOLORE_LISTINO*': '',
        '*SUPPLEMENTO_BICOLORE_SCONTATO*': '',
        '*TOTALE_RIGA_COLORE_TELAIO*': '',
        '*TOTALE_RIGA_COLORE_COPRIFILI*': '',
        '*SUPPLEMENTO_COLORE_INSERTO_LISTINO*': '', '*SUPPLEMENTO_COLORE_INSERTO_SCONTATO*': '',
        '*SUPPLEMENTO_COLORE_PIETRA*': '', '*SUPPLEMENTO_COLORE_PIETRA_SCONTATO*': '',
        '*PREZZO_VETRO_LISTINO*': fmt_euro(r.get('prezzo_vetro','')) if r.get('prezzo_vetro') else '',
        '*PREZZO_VETRO_SCONTATO*': s(r.get('prezzo_vetro',0)) if r.get('prezzo_vetro') else '',
        '*PREZZO_INCISIONE_O_STAMPA_LISTINO*': '', '*PREZZO_INCISIONE_O_STAMPA_SCONTATO*': '',
        '*SUPPLEMENTO_SERRATURA_LISTINO*': fmt_euro(r.get('prezzo_serratura','')) if r.get('prezzo_serratura') else '',
        '*SUPPLEMENTO_SERRATURA_SCONTATO*': s(r.get('prezzo_serratura',0)) if r.get('prezzo_serratura') else '',
        '*SUPPLEMENTO_STIPITE_LISTINO*': fmt_euro(r.get('prezzo_telaio','')) if r.get('prezzo_telaio') else '',
        '*SUPPLEMENTO_STIPITE_SCONTATO*': s(r.get('prezzo_telaio',0)) if r.get('prezzo_telaio') else '',
        '*SUPPLEMENTO_COLORE_FERRAMENTA_LISTINO*': fmt_euro(r.get('prezzo_ferramenta','')) if r.get('prezzo_ferramenta') else '',
        '*SUPPLEMENTO_COLORE_FERRAMENTA_SCONTATO*': s(r.get('prezzo_ferramenta',0)) if r.get('prezzo_ferramenta') else '',
        '*PREZZO_MANIGLIA_LISTINO*': fmt_euro(r.get('prezzo_maniglia','')) if r.get('prezzo_maniglia') else '',
        '*PREZZO_MANIGLIA_SCONTATO*': s(r.get('prezzo_maniglia',0)) if r.get('prezzo_maniglia') else '',
        '*PREZZO_KIT_VARSAVIA_SCORREVOLE_LISTINO*': '', '*PREZZO_KIT_VARSAVIA_SCORREVOLE_SCONTATO*': '',
        '*PREZZO_KIT_RIM_16_SCORREVOLE_LISTINO*': '', '*PREZZO_KIT_RIM_16_SCORREVOLE_SCONTATO*': '',
        '*PREZZO_LAVORAZIONI_EXTRA_LISTINO*': fmt_euro(r.get('prezzo_extra','')) if r.get('prezzo_extra') else '',
        '*PREZZO_LAVORAZIONI_EXTRA_SCONTATO*': '',
        '*PREZZO_ACCESSORI_LISTINO*': '', '*PREZZO_ACCESSORI_SCONTATO*': '',
        '*SUPPLEMENTO_FUORI_MISURA_L_LISTINO*': '', '*SUPPLEMENTO_FUORI_MISURA_L_SCONTATO*': '',
        '*SUPPLEMENTO_FUORI_MISURA_H_LISTINO*': '', '*SUPPLEMENTO_FUORI_MISURA_H_SCONTATO*': '',
        '*SUPPLEMENTO_RIFILATURA_LISTINO*': '', '*SUPPLEMENTO_RIFILATURA_SCONTATO*': '',
        '*TOTALE_RIGA*': '',
        '*TOTALE_RIGA_MODELLO*': s(r.get('prezzo_base', 0)),
        '*TOTALE_RIGA_COLORE*': s(r.get('prezzo_finitura',0)) if r.get('prezzo_finitura') else '',
        '*TOTALE_RIGA_COLORE_TELAIO*': '',
        '*TOTALE_RIGA_COLORE_COPRIFILI*': '',
        '*TOTALE_RIGA_COLORE_PIETRA*': '',
        '*TOTALE_RIGA_COLORE_INSERTO*': s(r.get('prezzo_ferramenta',0)) if r.get('prezzo_ferramenta') else '',
        '*TOTALE_RIGA_VETRO*': s(r.get('prezzo_vetro',0)) if r.get('prezzo_vetro') else '',
        '*TOTALE_RIGA_INCISIONE_O_STAMPA*': s(r.get('prezzo_extra',0)) if r.get('prezzo_extra') else '',
        '*TOTALE_RIGA_TIPOLOGIA*': '',
        '*TOTALE_RIGA_SERRATURA*': s(r.get('prezzo_serratura',0)) if r.get('prezzo_serratura') else '',
        '*TOTALE_RIGA_STIPITE*': s(r.get('prezzo_telaio',0)) if r.get('prezzo_telaio') else '',
        '*TOTALE_RIGA_COLORE_FERRAMENTA*': s(r.get('prezzo_ferramenta',0)) if r.get('prezzo_ferramenta') else '',
        '*TOTALE_RIGA_MANIGLIA*': s(r.get('prezzo_maniglia',0)) if r.get('prezzo_maniglia') else '',
        '*TOTALE_RIGA_KIT_VARSAVIA_SCORREVOLE*': '',
        '*TOTALE_RIGA_KIT_RIM_16_SCORREVOLE*': '',
        '*TOTALE_RIGA_LAVORAZIONI_EXTRA*': '',
        '*TOTALE_RIGA_ACCESSORI*': '',
        '*TOTALE_RIGA_FUORI_MISURA_L*': '',
        '*TOTALE_RIGA_FUORI_MISURA_H*': '',
        '*TOTALE_RIGA_RIFILATURA*': '',
        '*TOTALE_POSIZIONE*': fmt_euro(round(float(r.get('prezzo_unitario',0) or r.get('prezzo_base',0) or 0) * (1-sc/100) * int(r.get('quantita',1) or 1), 2)),
        '*IMMAGINE*': '',
    }

def xlsx_to_pdf(xlsx_path, out_dir):
    """Converti un xlsx in PDF e ritorna il path del PDF."""
    # LibreOffice necessita di una home dir scrivibile
    lo_home = '/tmp/lo_home'
    os.makedirs(lo_home, exist_ok=True)
    env = {**os.environ, 'HOME': lo_home, 'TMPDIR': '/tmp'}
    r = subprocess.run(
        ['libreoffice', '--headless', '--norestore', '--nofirststartwizard',
         '--convert-to', 'pdf', '--outdir', out_dir, xlsx_path],
        capture_output=True, text=True, timeout=60, env=env)
    pdf = xlsx_path.replace('.xlsx', '.pdf')
    if not os.path.exists(pdf):
        print(f"LO stderr: {r.stderr[:300]}", file=sys.stderr)
        print(f"LO stdout: {r.stdout[:300]}", file=sys.stderr)
    return pdf if os.path.exists(pdf) else None

def _leggi_footer_template(sheet_name):
    """Legge il blocco headerFooter XML originale dal template per lo sheet indicato."""
    import zipfile as _zft, re as _ret
    sheet_idx = {'PRIMA_PAGINA': 1, 'PAGINE_INTERMEDIE': 2, 'PAGINA_FINALE': 3}
    idx = sheet_idx.get(sheet_name, 1)
    sheet_file = f'xl/worksheets/sheet{idx}.xml'
    try:
        with _zft.ZipFile(TEMPLATE_PATH, 'r') as _zt:
            _orig_xml = _zt.read(sheet_file).decode('utf-8')
        _m = _ret.search(r'<headerFooter>.*?</headerFooter>', _orig_xml, _ret.DOTALL)
        return _m.group(0) if _m else None
    except Exception as _e:
        print(f'Footer template warning: {_e}', file=sys.stderr)
        return None

def genera_foglio(sheet_name, mapping, tmp_path, num_pagina=1, totale_pagine=1):
    """Carica il template, compila il foglio indicato, salva come xlsx."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[sheet_name]
    sostituisci(ws, mapping)
    # Pulisci pattern residui: sostituisci PRIMA con nome+versione combinati
    # poi pulisci cella per cella i trattini rimasti
    from openpyxl.cell.cell import MergedCell as _MC2
    _nome_man = mapping.get('*MANIGLIA*', '')
    _vers_man = mapping.get('*VERSIONE_MANIGLIA*', '')
    _man_display = _nome_man
    if _vers_man: _man_display = (_nome_man + ' - ' + _vers_man).strip(' -')
    # Sostituisci direttamente la cella combinata prima della pulizia generale
    for _row2 in ws.iter_rows():
        for _cell2 in _row2:
            if isinstance(_cell2, _MC2): continue
            if _cell2.value and isinstance(_cell2.value, str):
                _v2 = _cell2.value
                # Sostituisci il pattern combinato maniglia - versione
                if '*MANIGLIA*' in _v2 or '*VERSIONE_MANIGLIA*' in _v2:
                    _cell2.value = _man_display if _man_display else None
                    continue
                # Pulisci trattini residui da altre sostituzioni parziali
                _v2 = _v2.strip()
                if _v2 in ('-', ' - ', '- ', ' -'): _cell2.value = None; continue
                if _v2.endswith(' -') or _v2.endswith(' - '):
                    _v2 = _v2.rstrip(' -').rstrip()
                    _cell2.value = _v2 if _v2 else None
        # Sostituisci segnaposti nel footer — patch diretta XML dopo save
    # (openpyxl aggiunge tag vuoti che LibreOffice headless ignora)
    _footer_num = str(num_pagina)
    _footer_tot = str(totale_pagine)
    pulisci_euro_vuoti(ws, sheet_name)
    nascondi_vuote(ws, sheet_name)
    # Rimuovi gli altri fogli
    for s in list(wb.sheetnames):
        if s != sheet_name: del wb[s]
    wb.save(tmp_path)
    # Patch footer diretto nell'XML: openpyxl aggiunge tag vuoti che LibreOffice ignora
    # Sostituiamo il blocco headerFooter con versione pulita come nel template originale
    import zipfile as _zf, re as _re2
    _hf_template = _leggi_footer_template(sheet_name)
    _tmp_patch = tmp_path + '_fp'
    try:
        with _zf.ZipFile(tmp_path, 'r') as _zin:
            _items = [(_i, _zin.read(_i.filename)) for _i in _zin.infolist()]
        with _zf.ZipFile(_tmp_patch, 'w', _zf.ZIP_DEFLATED) as _zout:
            for _item, _data in _items:
                if _item.filename.startswith('xl/worksheets/sheet') and _item.filename.endswith('.xml'):
                    _xml = _data.decode('utf-8')
                    # Usa il footer letto dal template per il foglio corretto
                    if _hf_template:
                        _hf_new = _hf_template.replace('*NUMERO_PAGINA*', _footer_num).replace('*TOTALE_PAGINE*', _footer_tot)
                        _xml = _re2.sub(r'<headerFooter>.*?</headerFooter>', _hf_new, _xml, flags=_re2.DOTALL)
                        if '<headerFooter>' not in _xml:
                            _xml = _xml.replace('</worksheet>', _hf_new + '</worksheet>')
                    _zout.writestr(_item, _xml.encode('utf-8'))
                else:
                    _zout.writestr(_item, _data)
        os.replace(_tmp_patch, tmp_path)
    except Exception as _e:
        print(f"Footer patch warning: {_e}", file=sys.stderr)
        try: os.remove(_tmp_patch)
        except: pass
    # Ripristina media e drawing dal template (openpyxl le perde)
    _ripristina_media(tmp_path)
    return tmp_path

def _ripristina_media(xlsx_path):
    """Copia solo i rels dei worksheet dal template — NON le immagini."""
    import zipfile as zf2
    tmp_p = xlsx_path + '_tmp'
    try:
        with zf2.ZipFile(TEMPLATE_PATH, 'r') as zt:
            extra = {n: zt.read(n) for n in zt.namelist()
                     if n.startswith('xl/worksheets/_rels/')}

        with zf2.ZipFile(xlsx_path, 'r') as zin:
            items = [(i, zin.read(i.filename)) for i in zin.infolist()]
        existing = {i.filename for i, _ in items}

        with zf2.ZipFile(tmp_p, 'w', zf2.ZIP_DEFLATED) as zout:
            for item, data in items:
                zout.writestr(item, data)
            for name, data in extra.items():
                if name not in existing:
                    zout.writestr(name, data)

        os.replace(tmp_p, xlsx_path)
    except Exception as e:
        print(f"Media warning: {e}", file=sys.stderr)
        try: os.remove(tmp_p)
        except: pass

def genera_preventivo(dati_json, output_path):
    data = json.loads(dati_json)
    doc = data['documento']
    righe = data['righe']
    global _POS_RANGES
    _POS_RANGES = {}  # reset cache per ogni documento
    if not righe: print("Nessuna riga", file=sys.stderr); sys.exit(1)

    sc1 = float(doc.get('sconto1', 0))
    tot_imp = float(doc.get('totale_imponibile', 0))
    tot_netto = round(tot_imp*(1-sc1/100), 2)
    iva_euro = round(tot_netto*22/100, 2)
    tot_iva = round(tot_netto+iva_euro, 2)

    md = {
        '*TIPO_DOCUMENTO*': doc.get('tipo_documento','PREVENTIVO'),
        '*DATA_GENERAZIONE_DOCUMENTO': doc.get('data',''),
        '*CODICE_PREVENTIVO*': doc.get('numero',''),
        '*DATA_ULTIMA_MODIFICA*': doc.get('data_modifica',''),
        '*NOME_COGNOME_COMPILATORE*': doc.get('compilatore',''),
        '*RAGIONE SOCIALE*': doc.get('ragione_sociale',''),
        '*RAGIONE SOCIALE RIGA 2*': doc.get('ragione_sociale_riga2',''),
        '*INDIRIZZO*': doc.get('indirizzo',''),
        '*INDIRIZZO RIGA 2*': '',
        '*CAP* *CITTÀ* *(PROVINCIA)*': f"{doc.get('cap','')} {doc.get('citta','')} ({doc.get('provincia','')})",
        '*PAESE*': doc.get('paese','Italia'),
        '*NOME DESTINAZIONE*': doc.get('dest_nome', doc.get('ragione_sociale','')),
        '*NOME DESTINAZIONE RIGA 2*': '',
        '*INDIRIZZO DESTINAZIONE*': doc.get('dest_indirizzo', doc.get('indirizzo','')),
        '*CAP* *CITTÀ DEST* *(PROVINCIA DEST)*': f"{doc.get('dest_cap','')} {doc.get('dest_citta','')} ({doc.get('dest_provincia','')})",
        '*PAESE DEST*': doc.get('dest_paese','Italia'),
        '*RIFERIMENTO_CLIENTE*': doc.get('riferimento_cliente',''),
        '*CODICE_CLIENTE*': doc.get('codice_cliente',''),
        '*AGENTE*': doc.get('agente',''),
        '*RESA*': doc.get('resa',''),
        '*TRASPORTO*': doc.get('trasporto',''),
        '*PARTITA_IVA*': doc.get('partita_iva',''),
        '*CODICE_FISCALE*': doc.get('codice_fiscale',''),
        '*TELEFONO*': doc.get('telefono',''),
        '*CELLULARE*': doc.get('cellulare',''),
        '*REFERENTE*': doc.get('referente',''),
        '*RIFERIMENTI_DESTINAZIONE*': doc.get('dest_riferimenti',''),
        '*CONDIZIONI_DI_PAGAMENTO*': doc.get('condizioni_pagamento',''),
        '*BANCA_DI_APPOGGIO*': doc.get('banca',''),
        '*CIN* *ABI* *CAB*': doc.get('cin_abi_cab',''),
        '*CIN*': doc.get('cin',''),
        '*ABI*': doc.get('abi',''),
        '*CAB*': doc.get('cab',''),
        '*EMAIL_1*': doc.get('email1',''),
        '*EMAIL_ORDINI*': doc.get('email_ordini',''),
        '*EMAIL_2*': doc.get('email2',''),
        '*CODICE_SDI*': doc.get('sdi',''),
        '*SDI*': doc.get('sdi',''),
        '*PEC*': doc.get('pec',''),
        '*PEC_FATTURAZIONE*': doc.get('pec_fatturazione',''),
        '*REFERENTE*': doc.get('referente',''),
        '*TELEFONO_REFERENTE*': doc.get('telefono_referente',''),
        '*CELLULARE_REFERENTE*': doc.get('cellulare_referente',''),
        '*EMAIL_REFERENTE*': doc.get('email_referente',''),
        '*GIORNI_VALIDITÀ_OFFERTA*': str(doc.get('validita_offerta','30')),
        '*BARCODE_DOCUMENTO*': '',
        '*SOMMA_TOTALI_POSIZIONI*': fmt_euro(tot_imp),
        '*SCONTO*': fmt_euro(round(tot_imp - tot_netto, 2)) if sc1 else '',
        '*OMAGGI*': '', '*SCONTO_PAGAMENTO*': '',
        '*TOTALE_MERCE_SCONTATO*': fmt_euro(tot_netto),
        '*TOTALE_IMPONIBILE*': fmt_euro(tot_netto),
        '*TOTALE_IVA*': fmt_euro(iva_euro),
        '*TOTALE_IMBALLO*': fmt_euro(doc.get('totale_imballo','')),
        '*TOTALE_TRASPORTO*': '', '*TOTALE_SPESE*': '',
        '*SOMMA_RIEPILOGO_OFFERTA*': fmt_euro(tot_iva),
    }

    tmp_dir = os.path.dirname(os.path.abspath(output_path))
    base = os.path.basename(output_path).replace('.pdf','')
    xlsx_files = []
    pdf_files = []

    # PRIMA PAGINA
    totale_pagine = 1 + len(righe[1:]) + 1  # prima + intermedie + finale
    p = os.path.join(tmp_dir, f'{base}_p1.xlsx')
    genera_foglio('PRIMA_PAGINA', {**md, **map_riga(righe[0], sc1)}, p, num_pagina=1, totale_pagine=totale_pagine)
    xlsx_files.append(p)

    # PAGINE INTERMEDIE (pos 2..N)
    for i, riga in enumerate(righe[1:], 2):
        p = os.path.join(tmp_dir, f'{base}_p{i}.xlsx')
        genera_foglio('PAGINE_INTERMEDIE', {**md, **map_riga(riga, sc1)}, p, num_pagina=i, totale_pagine=totale_pagine)
        xlsx_files.append(p)

    # PAGINA FINALE
    p = os.path.join(tmp_dir, f'{base}_fin.xlsx')
    genera_foglio('PAGINA_FINALE', {**md}, p, num_pagina=totale_pagine, totale_pagine=totale_pagine)
    xlsx_files.append(p)

    # Converti ogni xlsx in PDF
    for xlsx in xlsx_files:
        pdf = xlsx_to_pdf(xlsx, tmp_dir)
        if pdf:
            pdf_files.append(pdf)
        else:
            print(f"Errore conversione: {xlsx}", file=sys.stderr)

    if not pdf_files:
        print("Nessun PDF generato", file=sys.stderr); sys.exit(1)

    # Unisci tutti i PDF
    if len(pdf_files) == 1:
        os.replace(pdf_files[0], output_path)
    else:
        # Prova pdftk
        r = subprocess.run(
            ['pdftk'] + pdf_files + ['cat', 'output', output_path],
            capture_output=True, text=True, timeout=60)
        if r.returncode != 0:
            # Fallback: ghostscript
            r2 = subprocess.run(
                ['gs','-dBATCH','-dNOPAUSE','-q','-sDEVICE=pdfwrite',
                 f'-sOutputFile={output_path}'] + pdf_files,
                capture_output=True, text=True, timeout=60)
            if r2.returncode != 0:
                # Ultimo fallback: usa solo la prima pagina
                import shutil
                shutil.copy(pdf_files[0], output_path)
                print(f"Warning: merge fallito, PDF parziale", file=sys.stderr)

    # Pulizia
    for f in xlsx_files + pdf_files:
        try: os.remove(f)
        except: pass

    # Aggiungi logo
    aggiungi_logo(output_path, LOGO_PATH)

    print(f"genera_pdf.py {VERSION} - template: {TEMPLATE_PATH}", file=sys.stderr)

if __name__ == '__main__':
    if len(sys.argv) > 1:
        with open(sys.argv[1],'r') as f: input_json = f.read()
    else:
        input_json = sys.stdin.read()
    output = sys.argv[2] if len(sys.argv) > 2 else '/tmp/output.pdf'
    genera_preventivo(input_json, output)
